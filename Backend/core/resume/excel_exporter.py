import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional
import os
from datetime import datetime


class ExcelExporter:
    def __init__(self):
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.alternate_row_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    
    def export_resumes(
        self,
        resumes: List[Dict[str, Any]],
        output_path: str,
        sheet_name: str = "Resumes"
    ) -> bool:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            headers = [
                "Name",
                "Email",
                "Mobile Number",
                "Education",
                "Skills",
                "Company Name",
                "College Name",
                "Designation",
                "Experience",
                "Total Experience (Years)",
                "Extracted Date"
            ]
            
            self._write_headers(ws, headers)
            
            for row_idx, resume in enumerate(resumes, start=2):
                self._write_resume_row(ws, row_idx, resume, headers)
            
            self._adjust_column_widths(ws, headers)
            
            wb.save(output_path)
            return True
        
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False
    
    def _write_headers(self, ws, headers: List[str]):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.cell_border
    
    def _write_resume_row(
        self,
        ws,
        row_idx: int,
        resume: Dict[str, Any],
        headers: List[str]
    ):
        mapping = {
            "Name": "name",
            "Email": "email",
            "Mobile Number": "phone",
            "Education": "education",
            "Skills": "skills",
            "Company Name": "company_name",
            "College Name": "college_name",
            "Designation": "designation",
            "Experience": "experience",
            "Total Experience (Years)": "total_experience",
            "Extracted Date": "extracted_date"
        }
        
        for col_idx, header in enumerate(headers, start=1):
            key = mapping.get(header, header.lower().replace(" ", "_"))
            value = ""
            
            if key == "extracted_date":
                value = resume.get("extracted_date", datetime.now().strftime("%Y-%m-%d"))
            elif key in ["education", "skills", "company_name", "college_name", "designation", "experience"]:
                data = resume.get(key, [])
                if isinstance(data, list):
                    value = ", ".join(str(item) for item in data)
                elif isinstance(data, str):
                    value = data
            else:
                value = str(resume.get(key, ""))
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = self.cell_border
            
            if row_idx % 2 == 0:
                cell.fill = self.alternate_row_fill
    
    def _adjust_column_widths(self, ws, headers: List[str]):
        column_widths = {
            "Name": 25,
            "Email": 30,
            "Mobile Number": 15,
            "Education": 40,
            "Skills": 35,
            "Company Name": 35,
            "College Name": 35,
            "Designation": 30,
            "Experience": 40,
            "Total Experience (Years)": 20,
            "Extracted Date": 15
        }
        
        for col_idx, header in enumerate(headers, start=1):
            width = column_widths.get(header, 20)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        ws.row_dimensions[1].height = 25
    
    def export_single_resume(
        self,
        resume: Dict[str, Any],
        output_path: str
    ) -> bool:
        return self.export_resumes([resume], output_path)
    
    def append_resume(
        self,
        resume: Dict[str, Any],
        output_path: str
    ) -> bool:
        try:
            if os.path.exists(output_path):
                wb = openpyxl.load_workbook(output_path)
                ws = wb.active
                start_row = ws.max_row + 1
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Resumes"
                start_row = 1
                
                headers = [
                    "Name", "Email", "Mobile Number", "Education", "Skills",
                    "Company Name", "College Name", "Designation", "Experience",
                    "Total Experience (Years)", "Extracted Date"
                ]
                self._write_headers(ws, headers)
                start_row = 2
            
            resume["extracted_date"] = datetime.now().strftime("%Y-%m-%d")
            
            headers = [
                "Name", "Email", "Mobile Number", "Education", "Skills",
                "Company Name", "College Name", "Designation", "Experience",
                "Total Experience (Years)", "Extracted Date"
            ]
            self._write_resume_row(ws, start_row, resume, headers)
            
            wb.save(output_path)
            return True
        
        except Exception as e:
            print(f"Error appending to Excel: {e}")
            return False
    
    def create_summary_sheet(
        self,
        resumes: List[Dict[str, Any]],
        output_path: str
    ) -> bool:
        try:
            wb = openpyxl.Workbook()
            
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            summary_headers = ["Total Resumes", "Unique Skills", "Unique Companies", "Date"]
            ws_summary.append(summary_headers)
            
            all_skills = set()
            all_companies = set()
            
            for resume in resumes:
                skills = resume.get("skills", [])
                if isinstance(skills, list):
                    all_skills.update(skills)
                
                companies = resume.get("company_name", [])
                if isinstance(companies, list):
                    all_companies.update(companies)
            
            ws_summary.append([
                len(resumes),
                len(all_skills),
                len(all_companies),
                datetime.now().strftime("%Y-%m-%d")
            ])
            
            for cell in ws_summary[1]:
                cell.font = self.header_font
                cell.fill = self.header_fill
            
            ws_details = wb.create_sheet("All Resumes")
            
            headers = [
                "Name", "Email", "Mobile Number", "Education", "Skills",
                "Company Name", "College Name", "Designation", "Experience",
                "Total Experience (Years)", "Extracted Date"
            ]
            
            self._write_headers(ws_details, headers)
            
            for row_idx, resume in enumerate(resumes, start=2):
                self._write_resume_row(ws_details, row_idx, resume, headers)
            
            wb.save(output_path)
            return True
        
        except Exception as e:
            print(f"Error creating summary: {e}")
            return False


_excel_exporter = None

def get_excel_exporter() -> ExcelExporter:
    global _excel_exporter
    if _excel_exporter is None:
        _excel_exporter = ExcelExporter()
    return _excel_exporter
