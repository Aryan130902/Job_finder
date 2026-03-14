"""
Excel Export Service for Resumes.

This module provides functionality to export resume data to Excel format.
"""

from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import Workbook


class ExcelExporter:
    """
    Service for exporting resume data to Excel format.
    
    Provides methods to export single resumes or batches to Excel
    with formatted styling.
    """
    
    def __init__(self):
        """Initialize the Excel exporter."""
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    def export_single_resume(
        self,
        resume_data: Dict[str, Any],
        output_path: str
    ) -> bool:
        """
        Export a single resume to Excel.
        
        Args:
            resume_data: Dictionary containing resume information
            output_path: Path to save the Excel file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Resume"
            
            headers = ["Field", "Value"]
            ws.append(headers)
            self._apply_header_style(ws, len(headers))
            
            self._add_row(ws, "Name", resume_data.get("name", ""))
            self._add_row(ws, "Email", resume_data.get("email", ""))
            self._add_row(ws, "Phone", resume_data.get("phone", ""))
            self._add_row(ws, "LinkedIn", resume_data.get("linkedin", ""))
            self._add_row(ws, "Portfolio", resume_data.get("portfolio", ""))
            
            skills = resume_data.get("skills", [])
            if skills:
                self._add_row(ws, "Skills", ", ".join(skills) if isinstance(skills, list) else skills)
            
            education = resume_data.get("education", [])
            if education:
                self._add_row(ws, "Education", ", ".join(education) if isinstance(education, list) else str(education))
            
            experience = resume_data.get("experience", [])
            if experience:
                self._add_row(ws, "Experience", ", ".join(experience) if isinstance(experience, list) else str(experience))
            
            company_name = resume_data.get("company_name", [])
            if company_name:
                self._add_row(ws, "Companies", ", ".join(company_name) if isinstance(company_name, list) else str(company_name))
            
            college_name = resume_data.get("college_name", [])
            if college_name:
                self._add_row(ws, "Colleges", ", ".join(college_name) if isinstance(college_name, list) else str(college_name))
            
            designation = resume_data.get("designation", [])
            if designation:
                self._add_row(ws, "Designations", ", ".join(designation) if isinstance(designation, list) else str(designation))
            
            self._add_row(ws, "Total Experience", resume_data.get("total_experience", ""))
            
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 60
            
            wb.save(output_path)
            return True
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False
    
    def export_resumes(
        self,
        resumes: List[Dict[str, Any]],
        output_path: str
    ) -> bool:
        """
        Export multiple resumes to Excel.
        
        Args:
            resumes: List of resume dictionaries
            output_path: Path to save the Excel file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Resumes"
            
            headers = [
                "Name", "Email", "Phone", "Skills", "Education",
                "Experience", "Companies", "Colleges", "Designations",
                "Total Experience"
            ]
            ws.append(headers)
            self._apply_header_style(ws, len(headers))
            
            for resume in resumes:
                skills = resume.get("skills", [])
                education = resume.get("education", [])
                experience = resume.get("experience", [])
                company_name = resume.get("company_name", [])
                college_name = resume.get("college_name", [])
                designation = resume.get("designation", [])
                
                ws.append([
                    resume.get("name", ""),
                    resume.get("email", ""),
                    resume.get("phone", ""),
                    ", ".join(skills) if isinstance(skills, list) else str(skills),
                    ", ".join(education) if isinstance(education, list) else str(education),
                    ", ".join(experience) if isinstance(experience, list) else str(experience),
                    ", ".join(company_name) if isinstance(company_name, list) else str(company_name),
                    ", ".join(college_name) if isinstance(college_name, list) else str(college_name),
                    ", ".join(designation) if isinstance(designation, list) else str(designation),
                    resume.get("total_experience", ""),
                ])
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(output_path)
            return True
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False
    
    def _apply_header_style(self, worksheet, num_cols: int):
        """Apply header styling to the worksheet."""
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.alignment
    
    def _add_row(self, worksheet, field: str, value: Any):
        """Add a field-value row to the worksheet."""
        worksheet.append([field, str(value) if value else ""])


_excel_exporter = None


def get_excel_exporter() -> ExcelExporter:
    """
    Get or create a singleton Excel exporter instance.
    
    Returns:
        ExcelExporter instance
    """
    global _excel_exporter
    if _excel_exporter is None:
        _excel_exporter = ExcelExporter()
    return _excel_exporter
